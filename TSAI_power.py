from tsai.all import *
import sklearn.metrics as skm
import openpyxl
import numpy as np

wb_new = openpyxl.load_workbook('preprocessedFiles/Reading_Test01-tfr.xlsx')
sheet_new = wb_new['Reading_Test01-tfr']

X_train = []
y_train = []
X_test = []
y_test = []

# epoch 0
for i in range(2, 20103):
  x_time = []
  for j in range(5, 66):
    _x = float(sheet_new.cell(row=i, column=j).value)
    x_time.append(_x)
  X_train.append(x_time)
  y = 1
  y_train.append(y)

# epoch 1
for i in range(20103, 40204):
  x_time = []
  for j in range(5, 66):
    _x = float(sheet_new.cell(row=i, column=j).value)
    x_time.append(_x)
  X_train.append(x_time)
  y = 0
  y_train.append(y)

# epoch 2
for i in range(40204, 60305):
  x_time = []
  for j in range(5, 66):
    _x = float(sheet_new.cell(row=i, column=j).value)
    x_time.append(_x)
  X_test.append(x_time)
  y = 0
  y_test.append(y)

# epoch 3
for i in range(60305, 80406):
  x_time = []
  for j in range(5, 66):
    _x = float(sheet_new.cell(row=i, column=j).value)
    x_time.append(_x)
  X_test.append(x_time)
  y = 1
  y_test.append(y)

X, y, splits = combine_split_data([np.array(X_train), np.array(X_test)], [np.array(y_train), np.array(y_test)])
X.shape, y.shape, splits

tfms  = [None, [Categorize()]]
dsets = TSDatasets(X, y, tfms=tfms, splits=splits, inplace=True)

dls = TSDataLoaders.from_dsets(dsets.train, dsets.valid, bs=[2,2], batch_tfms=[TSStandardize()], num_workers=0)

dls.show_batch(sharey=True)

model = InceptionTime(dls.vars, dls.c)
learn = Learner(dls, model, metrics=accuracy)
learn.save('stage0')

learn.load('stage0')
learn.lr_find()

learn.fit_one_cycle(2, lr_max=1e-4)
learn.save('stage1')

learn.recorder.plot_metrics()

learn.save_all(path='export', dls_fname='dls', model_fname='model', learner_fname='learner')

del learn, dsets, dls

learn = load_learner_all(path='export', dls_fname='dls', model_fname='model', learner_fname='learner')
dls = learn.dls
valid_dl = dls.valid
b = next(iter(valid_dl))

valid_probas, valid_targets, valid_preds = learn.get_preds(dl=valid_dl, with_decoded=True)

(valid_targets == valid_preds).float().mean()

learn.show_results()

learn.show_probas()

interp = ClassificationInterpretation.from_learner(learn)
interp.plot_confusion_matrix()

interp.most_confused(min_val=3)